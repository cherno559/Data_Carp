import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
from openpyxl import load_workbook
import io
import re

from predictor_module import render_predictor

# ── CONFIGURACIÓN DE LA PÁGINA ───────────────────────────────────────────────
st.set_page_config(
    page_title="Data CARP | River Plate Analytics",
    page_icon="🐔",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── SISTEMA DE DISEÑO: VARIABLES Y ESTILOS ───────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Rajdhani:wght@400;500;600;700&family=Inter:wght@300;400;500;600&display=swap');

:root {
    --red-primary:   #D0021B;
    --red-hover:     #A80016;
    --red-light:     rgba(208,2,27,0.08);
    --red-glow:      rgba(208,2,27,0.25);
    --white:         #FFFFFF;
    --gray-50:       #F9FAFB;
    --gray-100:      #F3F4F6;
    --gray-200:      #E5E7EB;
    --gray-400:      #9CA3AF;
    --gray-600:      #4B5563;
    --gray-800:      #1F2937;
    --gray-900:      #111827;
    --black:         #0A0A0A;
    --gold:          #C9A84C;
    --shadow-sm:     0 1px 3px rgba(0,0,0,0.08);
    --shadow-md:     0 4px 12px rgba(0,0,0,0.10);
    --shadow-lg:     0 8px 32px rgba(0,0,0,0.14);
    --radius-sm:     6px;
    --radius-md:     10px;
    --radius-lg:     16px;
}

html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
.main .block-container { padding: 1.5rem 2rem 3rem 2rem !important; max-width: 1400px; }

[data-testid="stSidebar"] { background: var(--black) !important; border-right: 3px solid var(--red-primary) !important; }
[data-testid="stSidebar"] * { color: var(--gray-200) !important; }
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stRadio label {
    color: var(--gray-400) !important; font-size: 11px !important;
    text-transform: uppercase !important; letter-spacing: 1.5px !important;
    font-weight: 600 !important; font-family: 'Rajdhani', sans-serif !important;
}
[data-testid="stSidebar"] .stRadio div[role="radio"] p {
    font-family: 'Rajdhani', sans-serif !important; font-weight: 600 !important;
    font-size: 15px !important; color: var(--gray-200) !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
    background: #1a1a1a !important; border: 1px solid #333 !important;
    color: var(--white) !important; border-radius: var(--radius-sm) !important;
}
[data-testid="stSidebar"] hr { border-color: #2a2a2a !important; margin: 1rem 0 !important; }

.sidebar-brand { display: flex; align-items: center; gap: 12px; padding: 4px 0 16px 0; }
.sidebar-brand-text .title { font-family: 'Bebas Neue', cursive; font-size: 32px; color: var(--white) !important; line-height: 1; letter-spacing: 2px; }
.sidebar-brand-text .subtitle { font-family: 'Rajdhani', sans-serif; font-size: 11px; color: var(--red-primary) !important; letter-spacing: 3px; text-transform: uppercase; font-weight: 700; }

.sidebar-section-label { font-family: 'Rajdhani', sans-serif !important; font-size: 10px !important; text-transform: uppercase !important; letter-spacing: 2px !important; color: #555 !important; font-weight: 700 !important; margin-bottom: 6px !important; margin-top: 8px !important; }

h1.page-title { font-family: 'Bebas Neue', cursive !important; font-size: 48px !important; color: var(--gray-900) !important; letter-spacing: 3px !important; margin: 0 0 4px 0 !important; line-height: 1 !important; }
.page-subtitle { font-family: 'Rajdhani', sans-serif; font-size: 13px; color: var(--gray-400); text-transform: uppercase; letter-spacing: 2px; font-weight: 600; margin-bottom: 24px; }
.section-title { font-family: 'Bebas Neue', cursive !important; font-size: 26px !important; color: var(--gray-800) !important; letter-spacing: 2px !important; margin: 0 0 12px 0 !important; }
.section-title-red { color: var(--red-primary) !important; }

.red-divider { height: 3px; background: linear-gradient(90deg, var(--red-primary), transparent); border: none; margin: 8px 0 24px 0; border-radius: 2px; }

.score-card { background: linear-gradient(135deg, var(--black) 0%, #1a1a1a 100%); border: 1px solid #2a2a2a; border-left: 4px solid var(--red-primary); border-radius: var(--radius-md); padding: 20px 28px; text-align: center; margin-bottom: 24px; box-shadow: var(--shadow-lg); position: relative; overflow: hidden; }
.score-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 1px; background: linear-gradient(90deg, var(--red-primary), transparent); }
.score-card .label { font-family: 'Rajdhani', sans-serif; font-size: 11px; color: var(--gray-400); text-transform: uppercase; letter-spacing: 3px; font-weight: 700; margin-bottom: 8px; }
.score-card .score { font-family: 'Bebas Neue', cursive; font-size: 52px; color: var(--white); line-height: 1; letter-spacing: 4px; }
.score-card .score .goals { color: var(--red-primary); }

[data-testid="stDataFrame"] { border-radius: var(--radius-md) !important; overflow: hidden !important; border: 1px solid var(--gray-200) !important; box-shadow: var(--shadow-sm) !important; }
[data-testid="stDataFrame"] table thead tr th { background: var(--black) !important; color: var(--white) !important; font-family: 'Rajdhani', sans-serif !important; font-size: 12px !important; text-transform: uppercase !important; letter-spacing: 1px !important; font-weight: 700 !important; padding: 10px 16px !important; }
[data-testid="stDataFrame"] table tbody tr:nth-child(even) td { background: var(--gray-50) !important; }
[data-testid="stDataFrame"] table tbody tr:hover td { background: var(--red-light) !important; }
[data-testid="stDataFrame"] table tbody tr td { font-family: 'Inter', sans-serif !important; font-size: 14px !important; padding: 9px 16px !important; color: var(--gray-800) !important; border-bottom: 1px solid var(--gray-100) !important; }

[data-testid="stMetric"] { background: var(--white) !important; border: 1px solid var(--gray-200) !important; border-radius: var(--radius-md) !important; padding: 16px !important; box-shadow: var(--shadow-sm) !important; }
[data-testid="stMetricValue"] { font-family: 'Bebas Neue', cursive !important; font-size: 32px !important; color: var(--red-primary) !important; }
[data-testid="stMetricLabel"] { font-family: 'Rajdhani', sans-serif !important; text-transform: uppercase !important; letter-spacing: 1px !important; font-size: 11px !important; font-weight: 700 !important; color: var(--gray-400) !important; }

.info-box { background: var(--red-light); border-left: 3px solid var(--red-primary); border-radius: 0 var(--radius-sm) var(--radius-sm) 0; padding: 10px 16px; margin-bottom: 16px; font-family: 'Rajdhani', sans-serif; font-size: 13px; color: var(--gray-600); font-weight: 500; }

.stTabs [data-baseweb="tab-list"] { gap: 4px; border-bottom: 2px solid var(--gray-200) !important; }
.stTabs [data-baseweb="tab"] { font-family: 'Rajdhani', sans-serif !important; font-weight: 700 !important; font-size: 14px !important; text-transform: uppercase !important; letter-spacing: 1px !important; color: var(--gray-400) !important; border-radius: var(--radius-sm) var(--radius-sm) 0 0 !important; padding: 10px 20px !important; }
.stTabs [aria-selected="true"] { color: var(--red-primary) !important; border-bottom: 3px solid var(--red-primary) !important; background: var(--red-light) !important; }

.player-header { display: flex; align-items: center; gap: 10px; padding: 14px 20px; border-radius: var(--radius-md); margin-bottom: 16px; font-family: 'Bebas Neue', cursive; font-size: 22px; letter-spacing: 2px; }
.player-header-red  { background: linear-gradient(135deg, var(--red-primary), #ff2d3a); color: white; }
.player-header-gray { background: linear-gradient(135deg, #374151, #6b7280); color: white; }

.footer { text-align: center; padding: 24px 0 8px 0; font-family: 'Rajdhani', sans-serif; font-size: 12px; color: var(--gray-400); letter-spacing: 1px; text-transform: uppercase; border-top: 1px solid var(--gray-200); margin-top: 48px; }
</style>
""", unsafe_allow_html=True)

# ── PALETA Y CONSTANTES ───────────────────────────────────────────────────────
COLORES_POSICION = {
    'DEF': '#3B82F6',
    'MED': '#22C55E',
    'DEL': '#EF4444',
    'POR': '#F59E0B'
}

PLOTLY_LAYOUT = dict(
    font=dict(family="Rajdhani, Inter, sans-serif", size=13, color="#1F2937"),
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(249,250,251,1)",
    margin=dict(l=16, r=16, t=40, b=16),
    title_font=dict(family="Bebas Neue, cursive", size=22, color="#1F2937"),
    legend=dict(
        bgcolor="rgba(255,255,255,0.9)",
        bordercolor="#E5E7EB",
        borderwidth=1,
        font=dict(family="Rajdhani", size=12),
    ),
    hoverlabel=dict(
        bgcolor="#111827",
        bordercolor="#D0021B",
        font_color="white",
        font_family="Rajdhani",
        font_size=13,
    ),
)

def apply_plotly_style(fig, title="", xaxis_title="", yaxis_title=""):
    fig.update_layout(**PLOTLY_LAYOUT, title=title)
    fig.update_xaxes(
        title_text=xaxis_title,
        gridcolor="#E5E7EB",
        linecolor="#E5E7EB",
        zeroline=True,
        zerolinecolor="#D1D5DB",
        zerolinewidth=1,
        title_font=dict(family="Rajdhani", size=12, color="#6B7280"),
    )
    fig.update_yaxes(
        title_text=yaxis_title,
        gridcolor="#E5E7EB",
        linecolor="#E5E7EB",
        zeroline=True,
        zerolinecolor="#D1D5DB",
        zerolinewidth=1,
        title_font=dict(family="Rajdhani", size=12, color="#6B7280"),
    )
    return fig

# ── RUTAS Y LOGOS ─────────────────────────────────────────────────────────────
CARPETA = Path(__file__).parent

RUTA_LOGO_ACTUAL = CARPETA / "logo_river_actual.png"
RUTA_LOGO_RETRO  = CARPETA / "logo_river_retro.png"
RUTA_LOGO_CARP   = CARPETA / "logo_carp.png"

# ── DETECCIÓN DE TEMPORADAS ───────────────────────────────────────────────────
archivos_disponibles = list(CARPETA.glob("Base_Datos_River_*.xlsx"))
temporadas_dict = {}
for archivo in archivos_disponibles:
    match = re.search(r"Base_Datos_River_(\d{4})\.xlsx", archivo.name)
    if match:
        temporadas_dict[match.group(1)] = archivo

anios_disponibles = sorted(list(temporadas_dict.keys()), reverse=True)

# ── CARGA DE DATOS ────────────────────────────────────────────────────────────
@st.cache_data
def cargar_datos_completos(ruta_excel):
    if not ruta_excel.exists():
        return pd.DataFrame(), f"❌ No se encontró el archivo: {ruta_excel.name}"
    try:
        xl = pd.ExcelFile(ruta_excel)
        partes = []
        hojas_omitir = ["Promedios Generales", "Goleadores", "Asistidores", "Resumen Estadísticas"]
        for hoja in xl.sheet_names:
            if hoja in hojas_omitir:
                continue
            df = pd.read_excel(ruta_excel, sheet_name=hoja)
            df.columns = df.columns.astype(str).str.strip()
            if 'Jugador' not in df.columns:
                continue
            df['Jugador'] = df['Jugador'].fillna("").astype(str).str.strip().str.title()
            df['Jugador'] = df['Jugador'].apply(lambda x: re.sub(r'\s+', ' ', str(x)))
            reemplazos = {'Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','á':'a','é':'e','í':'i','ó':'o','ú':'u'}
            for con_tilde, sin_tilde in reemplazos.items():
                df['Jugador'] = df['Jugador'].str.replace(con_tilde, sin_tilde, regex=False)
            if 'Minutos' in df.columns:
                df['Minutos'] = pd.to_numeric(df['Minutos'], errors="coerce").fillna(0)
            else:
                df['Minutos'] = 0
            df = df[(df['Jugador'] != "") & (df['Jugador'].str.lower() != "nan") & (df['Minutos'] > 0)]
            if 'Nota SofaScore' in df.columns:
                df['Nota SofaScore'] = pd.to_numeric(df['Nota SofaScore'], errors="coerce")
                df = df[~(df['Nota SofaScore'] == 0)]
            else:
                df['Nota SofaScore'] = np.nan
            cols_num = ['Goles','Asistencias','Pases Clave','Quites (Tackles)','Intercepciones','Tiros Totales','Efectividad Pases']
            for col in cols_num:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            df['Hoja_Original'] = hoja
            df['Partido'] = hoja
            partes.append(df)
        return pd.concat(partes, ignore_index=True) if partes else pd.DataFrame(), "OK"
    except Exception as e:
        return pd.DataFrame(), f"Error al cargar {ruta_excel.name}: {str(e)}"

@st.cache_data
def extraer_imagen_incrustada(ruta_excel_str, nombre_hoja, indice_imagen=0):
    try:
        wb = load_workbook(ruta_excel_str, data_only=True)
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            if hasattr(ws, '_images') and len(ws._images) > indice_imagen:
                img = ws._images[indice_imagen]
                return img._data() if callable(img._data) else img._data
        return None
    except Exception:
        return None

@st.cache_data
def extraer_estadisticas_equipo(ruta_excel_str, nombre_hoja):
    try:
        df = pd.read_excel(ruta_excel_str, sheet_name=nombre_hoja, header=None)
        row_idx, col_idx = None, None
        for r in range(min(120, len(df))):
            for c in range(min(15, len(df.columns))):
                val = str(df.iloc[r, c]).strip().lower()
                if val in ['métrica', 'metrica']:
                    row_idx, col_idx = r, c
                    break
            if row_idx is not None:
                break
        if row_idx is not None:
            df_team = df.iloc[row_idx+1:, col_idx:col_idx+3].copy()
            df_team.columns = df.iloc[row_idx, col_idx:col_idx+3].values
            df_team = df_team.dropna(subset=[df_team.columns[0]])
            df_team = df_team[df_team[df_team.columns[0]].astype(str).str.strip() != '']
            df_team = df_team.dropna(subset=[df_team.columns[1], df_team.columns[2]], how='all')
            return df_team
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()

@st.cache_data
def extraer_info_partido(ruta_excel_str, nombre_hoja):
    try:
        df = pd.read_excel(ruta_excel_str, sheet_name=nombre_hoja, header=None)
        local, rival = "Local", "Rival"
        g_local, g_rival = "?", "?"
        for r in range(min(120, len(df))):
            for c in range(min(15, len(df.columns))):
                val = str(df.iloc[r, c]).strip().lower()
                if val in ['métrica', 'metrica']:
                    local = str(df.iloc[r, c+1]).strip()
                    rival = str(df.iloc[r, c+2]).strip()
                if val == 'resultado':
                    g_local = str(df.iloc[r, c+1]).strip()
                    g_rival = str(df.iloc[r, c+2]).strip()
                    return local, rival, g_local, g_rival
        return local, rival, g_local, g_rival
    except Exception:
        return "Local", "Rival", "?", "?"

# BUG FIX 3: clean_goals definida fuera del loop para evitar closure bugs
def _clean_goals(g_str):
    m = re.match(r'^(\d+)', str(g_str).strip())
    return int(m.group(1)) if m else 0

@st.cache_data
def generar_historial_rivales(ruta_excel_str, hojas, condicion="Total"):
    historial = {}
    for hoja in hojas:
        local, rival, g_local, g_rival = extraer_info_partido(ruta_excel_str, hoja)
        if g_local == "?" or g_rival == "?":
            continue
        is_river_local = 'River' in local
        if condicion == "Local" and not is_river_local:
            continue
        if condicion == "Visitante" and is_river_local:
            continue
        gl = _clean_goals(g_local)
        gv = _clean_goals(g_rival)
        if is_river_local:
            equipo_rival = rival; gf = gl; gc = gv
        else:
            equipo_rival = local; gf = gv; gc = gl
        if equipo_rival not in historial:
            historial[equipo_rival] = {'PJ':0,'PG':0,'PE':0,'PP':0,'GF':0,'GC':0}
        historial[equipo_rival]['PJ'] += 1
        historial[equipo_rival]['GF'] += gf
        historial[equipo_rival]['GC'] += gc
        if gf > gc:
            historial[equipo_rival]['PG'] += 1
        elif gf < gc:
            historial[equipo_rival]['PP'] += 1
        else:
            historial[equipo_rival]['PE'] += 1
    df_hist = pd.DataFrame.from_dict(historial, orient='index').reset_index()
    if not df_hist.empty:
        df_hist.rename(columns={'index':'Rival'}, inplace=True)
        df_hist['DIF'] = df_hist['GF'] - df_hist['GC']
        df_hist = df_hist.sort_values(by=['PJ','PG','DIF'], ascending=[False,False,False])
    return df_hist

@st.cache_data
def generar_historial_completo(condicion="Total"):
    historial = {}
    for anio, ruta in temporadas_dict.items():
        try:
            xl = pd.ExcelFile(ruta)
            hojas_omitir = ["Promedios Generales","Goleadores","Asistidores","Resumen Estadísticas"]
            hojas_validas = [h for h in xl.sheet_names if h not in hojas_omitir]
            for hoja in hojas_validas:
                local, rival, g_local, g_rival = extraer_info_partido(str(ruta), hoja)
                if g_local == "?" or g_rival == "?":
                    continue
                is_river_local = 'River' in local
                if condicion == "Local" and not is_river_local:
                    continue
                if condicion == "Visitante" and is_river_local:
                    continue
                gl = _clean_goals(g_local)
                gv = _clean_goals(g_rival)
                if is_river_local:
                    equipo_rival = rival; gf = gl; gc = gv
                else:
                    equipo_rival = local; gf = gv; gc = gl
                if equipo_rival not in historial:
                    historial[equipo_rival] = {'PJ':0,'PG':0,'PE':0,'PP':0,'GF':0,'GC':0}
                historial[equipo_rival]['PJ'] += 1
                historial[equipo_rival]['GF'] += gf
                historial[equipo_rival]['GC'] += gc
                if gf > gc:
                    historial[equipo_rival]['PG'] += 1
                elif gf < gc:
                    historial[equipo_rival]['PP'] += 1
                else:
                    historial[equipo_rival]['PE'] += 1
        except Exception:
            continue
    df_hist = pd.DataFrame.from_dict(historial, orient='index').reset_index()
    if not df_hist.empty:
        df_hist.rename(columns={'index':'Rival'}, inplace=True)
        df_hist['DIF'] = df_hist['GF'] - df_hist['GC']
        df_hist = df_hist.sort_values(by=['PJ','PG','DIF'], ascending=[False,False,False])
    return df_hist

def mostrar_marcador(ruta_excel, hoja_excel):
    local, rival, g_local, g_rival = extraer_info_partido(str(ruta_excel), hoja_excel)
    if g_local != "?" and g_rival != "?":
        st.markdown(f"""
        <div class="score-card">
            <div class="label">Resultado Final</div>
            <div class="score">{local} <span class="goals">{g_local} – {g_rival}</span> {rival}</div>
        </div>
        """, unsafe_allow_html=True)

def page_header(icon, title, subtitle=""):
    st.markdown(f"""
    <div style="margin-bottom:4px;">
        <div style="font-family:'Bebas Neue',cursive;font-size:42px;color:#111827;letter-spacing:3px;line-height:1;">
            <span style="color:#D0021B;">{icon}</span> {title}
        </div>
        {"<div style='font-family:Rajdhani,sans-serif;font-size:12px;color:#9CA3AF;text-transform:uppercase;letter-spacing:2px;font-weight:600;margin-top:2px;'>" + subtitle + "</div>" if subtitle else ""}
    </div>
    <div class="red-divider"></div>
    """, unsafe_allow_html=True)

# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    col_logo, col_text = st.columns([1, 2])
    with col_logo:
        if RUTA_LOGO_ACTUAL.exists():
            st.image(str(RUTA_LOGO_ACTUAL), width=64)
    with col_text:
        st.markdown("""
        <div style="padding-top:6px;">
            <div style="font-family:'Bebas Neue',cursive;font-size:30px;color:white;line-height:1;letter-spacing:2px;">DATA CARP</div>
            <div style="font-family:'Rajdhani',sans-serif;font-size:10px;color:#D0021B;letter-spacing:3px;font-weight:700;text-transform:uppercase;">Analytics</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<hr style='border-color:#222;margin:12px 0;'>", unsafe_allow_html=True)

    if not anios_disponibles:
        st.error("❌ No se encontraron archivos de Base de Datos.")
        st.stop()

    st.markdown("<div class='sidebar-section-label'>Temporada</div>", unsafe_allow_html=True)
    temporada_sel = st.selectbox("", anios_disponibles, label_visibility="collapsed")
    EXCEL_ACTUAL = temporadas_dict[temporada_sel]

    st.markdown("<hr style='border-color:#222;margin:12px 0;'>", unsafe_allow_html=True)

    st.markdown("<div class='sidebar-section-label'>Categoría</div>", unsafe_allow_html=True)
    categoria = st.radio(
        "",
        ["🏆 Por Temporada", "🗓️ Por Fecha", "🔧 Herramientas"],
        label_visibility="collapsed"
    )

    st.markdown("<hr style='border-color:#222;margin:12px 0;'>", unsafe_allow_html=True)

    if categoria == "🏆 Por Temporada":
        st.markdown("<div class='sidebar-section-label'>Sección</div>", unsafe_allow_html=True)
        menu = st.radio("", ["Resumen General", "Historial", "Mapas de Rendimiento", "Análisis Individual"], label_visibility="collapsed")
    elif categoria == "🗓️ Por Fecha":
        st.markdown("<div class='sidebar-section-label'>Sección</div>", unsafe_allow_html=True)
        menu = st.radio("", ["Estadísticas de Equipo", "Estadísticas Individuales", "Parado Táctico", "Mapa de Tiros"], label_visibility="collapsed")
    elif categoria == "🔧 Herramientas":
        st.markdown("<div class='sidebar-section-label'>Sección</div>", unsafe_allow_html=True)
        menu = st.radio("", ["Predictor de Partidos", "Cara a Cara", "Historial General"], label_visibility="collapsed")

    st.markdown("<br><br>", unsafe_allow_html=True)

    col_b1, col_b2 = st.columns(2)
    with col_b1:
        if RUTA_LOGO_RETRO.exists():
            st.image(str(RUTA_LOGO_RETRO), width=72)
    with col_b2:
        if RUTA_LOGO_CARP.exists():
            st.image(str(RUTA_LOGO_CARP), width=72)

    st.markdown("""
    <div style="font-family:'Rajdhani',sans-serif;font-size:10px;color:#444;text-align:center;letter-spacing:1px;margin-top:8px;">
        DATA CARP · CLUB ATLÉTICO RIVER PLATE
    </div>
    """, unsafe_allow_html=True)

# ── PREDICTOR ─────────────────────────────────────────────────────────────────
if menu == "Predictor de Partidos":
    page_header("🤖", "PREDICTOR DE PARTIDOS", f"Poisson + Montecarlo · Temporada {temporada_sel}")
    render_predictor(EXCEL_ACTUAL, apply_plotly_style_fn=apply_plotly_style)
    st.markdown("<div class='footer'>Data CARP · Club Atlético River Plate · Análisis de Rendimiento</div>", unsafe_allow_html=True)
    st.stop()

# ── CARGA Y PROCESAMIENTO ─────────────────────────────────────────────────────
df_raw, estado = cargar_datos_completos(EXCEL_ACTUAL)
if estado != "OK":
    st.error(estado)
    st.stop()

if 'Efectividad Pases' in df_raw.columns:
    df_raw['Efectividad Pases'] = df_raw['Efectividad Pases'].replace(0, np.nan)

if 'Posición' in df_raw.columns:
    posiciones = df_raw.groupby('Jugador')['Posición'].agg(
        lambda x: x.mode()[0] if not x.mode().empty else '—'
    ).reset_index()
else:
    posiciones = pd.DataFrame({'Jugador': df_raw['Jugador'].unique(), 'Posición': '—'})

# BUG FIX 1 & 6: agg de Quites separado para evitar if/else en dict
_tiene_quites = 'Quites (Tackles)' in df_raw.columns
_col_quites = 'Quites (Tackles)' if _tiene_quites else ('Quites' if 'Quites' in df_raw.columns else None)

agg_dict = dict(
    Partidos=('Partido', 'nunique'),
    Promedio=('Nota SofaScore', 'mean'),
    Minutos=('Minutos', 'sum'),
    Goles=('Goles', 'sum'),
    Asistencias=('Asistencias', 'sum'),
    Intercepciones=('Intercepciones', 'sum'),
)
if 'Pases Clave' in df_raw.columns:
    agg_dict['Pases_Clave'] = ('Pases Clave', 'sum')
if 'Tiros Totales' in df_raw.columns:
    agg_dict['Tiros_Totales'] = ('Tiros Totales', 'sum')
if 'Efectividad Pases' in df_raw.columns:
    agg_dict['Efectividad_Pases'] = ('Efectividad Pases', 'mean')
if _col_quites:
    agg_dict['Quites_agg'] = (_col_quites, 'sum')

df_agrupado = df_raw.groupby('Jugador', as_index=False).agg(**agg_dict)

# Renombrar Quites a nombre estándar
if 'Quites_agg' in df_agrupado.columns:
    df_agrupado.rename(columns={'Quites_agg': 'Quites (Tackles)'}, inplace=True)
else:
    df_agrupado['Quites (Tackles)'] = 0

# Columnas opcionales que pueden no existir
for col_opt in ['Pases_Clave', 'Tiros_Totales', 'Efectividad_Pases']:
    if col_opt not in df_agrupado.columns:
        df_agrupado[col_opt] = 0

df_agrupado = df_agrupado.merge(posiciones, on='Jugador')
df_agrupado['Promedio'] = df_agrupado['Promedio'].round(2)
df_agrupado['Efectividad_Pases'] = df_agrupado['Efectividad_Pases'].round(1).fillna(0)

df_forma = df_raw.dropna(subset=['Nota SofaScore']).groupby('Jugador')['Nota SofaScore'].apply(
    lambda x: " | ".join([f"{n:.1f}" for n in list(x)[-5:]])
).reset_index(name='Forma (Últ. 5)')
df_agrupado = df_agrupado.merge(df_forma, on='Jugador', how='left')
df_agrupado['Forma (Últ. 5)'] = df_agrupado['Forma (Últ. 5)'].fillna('—')

total_partidos_temporada = df_raw['Partido'].nunique()

# =============================================================================
# PÁGINAS
# =============================================================================

# ─── RESUMEN GENERAL ──────────────────────────────────────────────────────────
if menu == "Resumen General":
    page_header("🐔", "PANEL GENERAL", f"Temporada {temporada_sel}")

    total_partidos  = total_partidos_temporada
    promedio_equipo = df_raw['Nota SofaScore'].mean()
    total_goles     = df_agrupado['Goles'].sum()
    total_asist     = df_agrupado['Asistencias'].sum()

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Partidos Analizados", int(total_partidos))
    k2.metric("Promedio SofaScore Equipo", f"{promedio_equipo:.2f}")
    k3.metric("Goles Totales", int(total_goles))
    k4.metric("Asistencias Totales", int(total_asist))

    st.markdown("<br>", unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["📊 Promedios SofaScore", "⚽ Goles & Asistencias", "📋 Plantel Completo"])

    with tab1:
        df_promedios = df_agrupado.dropna(subset=['Promedio'])[
            ['Jugador','Posición','Promedio','Partidos','Forma (Últ. 5)']
        ].sort_values('Promedio', ascending=False).reset_index(drop=True)
        df_promedios.index = df_promedios.index + 1

        if not df_promedios.empty:
            fig_prom = go.Figure()
            colores_barras = [COLORES_POSICION.get(p, '#9CA3AF') for p in df_promedios['Posición']]
            fig_prom.add_trace(go.Bar(
                x=df_promedios['Promedio'],
                y=df_promedios['Jugador'],
                orientation='h',
                marker=dict(color=colores_barras, line=dict(width=0)),
                text=df_promedios['Promedio'].apply(lambda x: f"{x:.2f}"),
                textposition='outside',
                textfont=dict(family="Rajdhani", size=12, color="#374151"),
                hovertemplate="<b>%{y}</b><br>Nota: %{x:.2f}<br>Partidos: %{customdata}<extra></extra>",
                customdata=df_promedios['Partidos'],
            ))
            fig_prom.add_vline(
                x=promedio_equipo,
                line_dash="dot",
                line_color="#D0021B",
                line_width=2,
                annotation_text=f"Prom. Equipo: {promedio_equipo:.2f}",
                annotation_position="top right",
                annotation_font=dict(family="Rajdhani", size=12, color="#D0021B"),
            )
            apply_plotly_style(fig_prom, xaxis_title="Nota Promedio SofaScore", yaxis_title="")
            fig_prom.update_layout(
                height=max(400, len(df_promedios) * 28),
                xaxis=dict(range=[5, 8.5]),
                yaxis=dict(categoryorder='total ascending'),
            )
            st.plotly_chart(fig_prom, use_container_width=True)
            st.markdown("<div class='info-box'>💡 La línea punteada roja indica el promedio del equipo. Los colores representan la posición: <b>🔵 DEF · 🟢 MED · 🔴 DEL · 🟠 POR</b></div>", unsafe_allow_html=True)

    with tab2:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("<div class='section-title'>⚽ GOLEADORES</div>", unsafe_allow_html=True)
            df_gol = df_agrupado[df_agrupado['Goles'] > 0][['Jugador','Posición','Goles']].sort_values('Goles', ascending=False).reset_index(drop=True)
            df_gol.index = df_gol.index + 1
            if not df_gol.empty:
                fig_gol = go.Figure(go.Bar(
                    x=df_gol['Jugador'], y=df_gol['Goles'],
                    marker_color='#D0021B',
                    text=df_gol['Goles'], textposition='outside',
                    textfont=dict(family="Bebas Neue", size=18, color="#D0021B"),
                    hovertemplate="<b>%{x}</b><br>Goles: %{y}<extra></extra>",
                ))
                apply_plotly_style(fig_gol, yaxis_title="Goles")
                fig_gol.update_layout(height=320, xaxis_tickangle=-25)
                st.plotly_chart(fig_gol, use_container_width=True)
            st.dataframe(df_gol, hide_index=False, use_container_width=True)

        with c2:
            st.markdown("<div class='section-title'>👟 ASISTIDORES</div>", unsafe_allow_html=True)
            df_ast = df_agrupado[df_agrupado['Asistencias'] > 0][['Jugador','Posición','Asistencias']].sort_values('Asistencias', ascending=False).reset_index(drop=True)
            df_ast.index = df_ast.index + 1
            if not df_ast.empty:
                fig_ast = go.Figure(go.Bar(
                    x=df_ast['Jugador'], y=df_ast['Asistencias'],
                    marker_color='#3B82F6',
                    text=df_ast['Asistencias'], textposition='outside',
                    textfont=dict(family="Bebas Neue", size=18, color="#3B82F6"),
                    hovertemplate="<b>%{x}</b><br>Asistencias: %{y}<extra></extra>",
                ))
                apply_plotly_style(fig_ast, yaxis_title="Asistencias")
                fig_ast.update_layout(height=320, xaxis_tickangle=-25)
                st.plotly_chart(fig_ast, use_container_width=True)
            st.dataframe(df_ast, hide_index=False, use_container_width=True)

    with tab3:
        cols_mostrar = ['Jugador','Posición','Partidos','Promedio','Minutos','Goles','Asistencias',
                        'Pases_Clave','Quites (Tackles)','Intercepciones','Forma (Últ. 5)']
        cols_existentes = [c for c in cols_mostrar if c in df_agrupado.columns]
        st.dataframe(
            df_agrupado[cols_existentes].sort_values('Minutos', ascending=False).reset_index(drop=True),
            hide_index=True, use_container_width=True, height=500,
        )

# ─── HISTORIAL ────────────────────────────────────────────────────────────────
elif menu == "Historial":
    page_header("📖", "HISTORIAL VS RIVALES", f"Temporada {temporada_sel}")
    condicion_sel = st.radio("Condición:", ["Total", "Local", "Visitante"], horizontal=True)
    st.markdown("<br>", unsafe_allow_html=True)
    hojas_unicas = df_raw.drop_duplicates('Partido')['Hoja_Original'].tolist()
    df_historial = generar_historial_rivales(str(EXCEL_ACTUAL), hojas_unicas, condicion_sel)
    if df_historial.empty:
        st.info(f"No hay partidos registrados bajo la condición '{condicion_sel}' en esta temporada.")
    else:
        total_pj = df_historial['PJ'].sum(); total_pg = df_historial['PG'].sum()
        total_pe = df_historial['PE'].sum(); total_pp = df_historial['PP'].sum()
        efectividad = (total_pg * 3 + total_pe) / (total_pj * 3) * 100 if total_pj > 0 else 0
        c1,c2,c3,c4,c5 = st.columns(5)
        c1.metric("Partidos Jugados", int(total_pj)); c2.metric("Victorias", int(total_pg))
        c3.metric("Empates", int(total_pe)); c4.metric("Derrotas", int(total_pp))
        c5.metric("Efectividad Puntos", f"{efectividad:.1f}%")
        st.markdown("<br>", unsafe_allow_html=True)
        c_graf, c_tabla = st.columns([1.5, 1])
        with c_graf:
            st.markdown("<div class='section-title'>📊 RESULTADOS POR RIVAL</div>", unsafe_allow_html=True)
            df_hist_graf = df_historial.sort_values(by='PJ', ascending=True)
            fig_hist = go.Figure()
            fig_hist.add_trace(go.Bar(y=df_hist_graf['Rival'], x=df_hist_graf['PG'], name='Ganados', orientation='h', marker=dict(color='#22C55E'), text=df_hist_graf['PG'].replace(0,''), textposition='inside', textfont=dict(color='white', family='Bebas Neue')))
            fig_hist.add_trace(go.Bar(y=df_hist_graf['Rival'], x=df_hist_graf['PE'], name='Empatados', orientation='h', marker=dict(color='#9CA3AF'), text=df_hist_graf['PE'].replace(0,''), textposition='inside', textfont=dict(color='white', family='Bebas Neue')))
            fig_hist.add_trace(go.Bar(y=df_hist_graf['Rival'], x=df_hist_graf['PP'], name='Perdidos', orientation='h', marker=dict(color='#EF4444'), text=df_hist_graf['PP'].replace(0,''), textposition='inside', textfont=dict(color='white', family='Bebas Neue')))
            apply_plotly_style(fig_hist, xaxis_title="Cantidad de Partidos", yaxis_title="")
            fig_hist.update_layout(barmode='stack', height=max(400, len(df_hist_graf)*35), legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig_hist, use_container_width=True)
            st.markdown("<div class='info-box'>💡 Los partidos definidos por penales se contabilizan estadísticamente como <b>Empate</b>.</div>", unsafe_allow_html=True)
        with c_tabla:
            st.markdown("<div class='section-title'>📋 TABLA DETALLADA</div>", unsafe_allow_html=True)
            st.dataframe(df_historial[['Rival','PJ','PG','PE','PP','GF','GC','DIF']], hide_index=True, use_container_width=True, height=600)

# ─── MAPAS DE RENDIMIENTO ─────────────────────────────────────────────────────
elif menu == "Mapas de Rendimiento":
    page_header("🗺️", "MAPAS DE RENDIMIENTO", f"Temporada {temporada_sel} · Estadísticas P90")
    with st.sidebar:
        st.markdown("<div class='sidebar-section-label'>Filtro de Minutos</div>", unsafe_allow_html=True)
        min_min = st.slider("Minutos mínimos", 0, int(df_agrupado['Minutos'].max()), 180, label_visibility="collapsed")
    df_p90 = df_agrupado[df_agrupado['Minutos'] >= min_min].copy()
    if df_p90.empty:
        st.warning("No hay jugadores con esa cantidad de minutos. Reducí el filtro.")
        st.stop()
    for col_base, col_p90 in [('Pases_Clave','PasesClave_P90'),('Asistencias','Asistencias_P90'),('Quites (Tackles)','Quites_P90'),('Intercepciones','Inter_P90'),('Goles','Goles_P90')]:
        if col_base in df_p90.columns:
            df_p90[col_p90] = (df_p90[col_base] / df_p90['Minutos'].replace(0,1)) * 90
    def scatter_cuadrantes(df, x_col, y_col, x_label, y_label, title):
        if x_col not in df.columns or y_col not in df.columns:
            return go.Figure()
        x_mean = df[x_col].mean(); y_mean = df[y_col].mean()
        fig = px.scatter(df, x=x_col, y=y_col, color='Posición', hover_name='Jugador', color_discrete_map=COLORES_POSICION, size_max=18)
        fig.update_traces(marker=dict(size=14, line=dict(width=1.5, color='white')), hovertemplate="<b>%{hovertext}</b><br>"+x_label+": %{x:.2f}<br>"+y_label+": %{y:.2f}<extra></extra>")
        fig.add_vline(x=x_mean, line_dash="dash", line_color="#9CA3AF", line_width=1, annotation_text=f"Prom. {x_label}: {x_mean:.2f}", annotation_font=dict(family="Rajdhani", size=11, color="#6B7280"), annotation_position="top right")
        fig.add_hline(y=y_mean, line_dash="dash", line_color="#9CA3AF", line_width=1, annotation_text=f"Prom. {y_label}: {y_mean:.2f}", annotation_font=dict(family="Rajdhani", size=11, color="#6B7280"), annotation_position="right")
        apply_plotly_style(fig, title=title, xaxis_title=x_label, yaxis_title=y_label)
        fig.update_layout(height=420, legend_title_text="Posición")
        return fig
    tab_def, tab_cre, tab_ata = st.tabs(["🛡️ Defensa", "🧠 Creación", "🎯 Ataque"])
    with tab_def:
        st.markdown("<div class='info-box'>Los jugadores en el cuadrante superior derecho son los más activos en recuperación de balón.</div>", unsafe_allow_html=True)
        fig_def = scatter_cuadrantes(df_p90, 'Quites_P90', 'Inter_P90', 'Quites por 90 min', 'Intercepciones por 90 min', 'PERFIL DEFENSIVO')
        st.plotly_chart(fig_def, use_container_width=True)
    with tab_cre:
        c1, c2 = st.columns(2)
        with c1:
            fig_kp = scatter_cuadrantes(df_p90, 'PasesClave_P90', 'Asistencias_P90', 'Pases Clave por 90 min', 'Asistencias por 90 min', 'CREACIÓN DE JUEGO')
            st.plotly_chart(fig_kp, use_container_width=True)
        with c2:
            df_pases = df_p90[df_p90['Efectividad_Pases'] > 0].copy()
            if not df_pases.empty:
                fig_ef = scatter_cuadrantes(df_pases, 'PasesClave_P90', 'Efectividad_Pases', 'Pases Clave por 90 min', 'Efectividad de Pases (%)', 'CALIDAD DE PASE')
                st.plotly_chart(fig_ef, use_container_width=True)
            else:
                st.info("Sin datos de efectividad de pases.")
    with tab_ata:
        df_t = df_agrupado[df_agrupado['Tiros_Totales'] > 0].copy()
        if not df_t.empty:
            df_t['Conv_Rate'] = (df_t['Goles'] / df_t['Tiros_Totales'] * 100).round(1)
            fig_of = px.scatter(df_t, x='Tiros_Totales', y='Goles', color='Posición', hover_name='Jugador', color_discrete_map=COLORES_POSICION, custom_data=['Conv_Rate','Minutos'])
            fig_of.update_traces(marker=dict(size=14, line=dict(width=1.5, color='white')), hovertemplate="<b>%{hovertext}</b><br>Tiros: %{x}<br>Goles: %{y}<br>Conversión: %{customdata[0]:.1f}%<extra></extra>")
            max_tiros = df_t['Tiros_Totales'].max()
            fig_of.add_shape(type="line", x0=0, y0=0, x1=max_tiros+2, y1=(max_tiros+2)*0.20, line=dict(color="#6B7280", dash="dot", width=1.5))
            fig_of.add_annotation(x=max_tiros+1, y=(max_tiros+1)*0.20, text="Ref. 20% conversión", font=dict(family="Rajdhani", size=11, color="#6B7280"), showarrow=False)
            apply_plotly_style(fig_of, title='EFICIENCIA GOLEADORA', xaxis_title='Tiros Totales', yaxis_title='Goles')
            fig_of.update_layout(height=420)
            st.plotly_chart(fig_of, use_container_width=True)
            st.markdown("<div class='info-box'>La línea punteada representa el umbral de referencia del 20% de conversión.</div>", unsafe_allow_html=True)
        else:
            st.info("Sin datos de tiros disponibles.")

# ─── ANÁLISIS INDIVIDUAL ──────────────────────────────────────────────────────
elif menu == "Análisis Individual":
    page_header("🔎", "ANÁLISIS INDIVIDUAL", f"Temporada {temporada_sel}")
    jugadores_ordenados = sorted(df_raw['Jugador'].unique())
    jugador_sel = st.selectbox("Seleccioná un jugador:", jugadores_ordenados)
    if jugador_sel:
        df_j = df_raw[df_raw['Jugador'] == jugador_sel].copy()
        pos = df_j['Posición'].mode()[0] if 'Posición' in df_j.columns and not df_j['Posición'].isnull().all() else "—"
        min_tot = int(df_j['Minutos'].sum()); prom = df_j['Nota SofaScore'].mean(); parts = df_j['Partido'].nunique()
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#111827,#1F2937);border-left:4px solid #D0021B;border-radius:10px;padding:20px 28px;margin-bottom:24px;display:flex;align-items:center;gap:20px;">
            <div>
                <div style="font-family:'Bebas Neue',cursive;font-size:36px;color:white;letter-spacing:2px;">{jugador_sel}</div>
                <div style="font-family:'Rajdhani',sans-serif;font-size:13px;color:#D0021B;letter-spacing:2px;font-weight:700;">{pos} · {parts} PARTIDOS · {min_tot} MINUTOS</div>
            </div>
            <div style="margin-left:auto;text-align:right;">
                <div style="font-family:'Bebas Neue',cursive;font-size:56px;color:#D0021B;line-height:1;">{prom:.2f}</div>
                <div style="font-family:'Rajdhani',sans-serif;font-size:11px;color:#9CA3AF;letter-spacing:2px;">NOTA PROMEDIO</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<div class='section-title'>📈 EVOLUCIÓN DE NOTAS</div>", unsafe_allow_html=True)
        df_j_notas = df_j.dropna(subset=['Nota SofaScore'])
        if not df_j_notas.empty:
            color_barras = ['#D0021B' if n >= prom else '#9CA3AF' for n in df_j_notas['Nota SofaScore']]
            fig_l = go.Figure()
            fig_l.add_trace(go.Bar(x=df_j_notas['Partido'], y=df_j_notas['Nota SofaScore'], marker_color=color_barras, text=df_j_notas['Nota SofaScore'].apply(lambda n: f"{n:.1f}"), textposition="outside", textfont=dict(family="Bebas Neue", size=16), hovertemplate="<b>%{x}</b><br>Nota: %{y:.1f}<extra></extra>"))
            fig_l.add_hline(y=prom, line_dash="dot", line_color="#D0021B", line_width=2, annotation_text=f"Promedio: {prom:.2f}", annotation_font=dict(family="Rajdhani", size=12, color="#D0021B"), annotation_position="top right")
            apply_plotly_style(fig_l, yaxis_title="Nota SofaScore")
            fig_l.update_layout(yaxis=dict(range=[4, 10.5]), xaxis=dict(tickangle=-30), height=340)
            st.plotly_chart(fig_l, use_container_width=True)
        st.markdown("<br>", unsafe_allow_html=True)
        c_radar, c_metrics = st.columns([1.6, 1])
        with c_radar:
            st.markdown("<div class='section-title'>🛡️ PERFIL TÁCTICO</div>", unsafe_allow_html=True)
            metrics_radar = ['Goles','Asistencias','Pases Clave','Quites (Tackles)','Intercepciones']
            labels_radar  = ['Goles','Asistencias','Pases Clave','Quites','Intercep.']
            totales_jugador = [df_j[m].sum() if m in df_j.columns else 0 for m in metrics_radar]
            df_squad_totals = df_raw.groupby('Jugador').agg(Minutos_tot=('Minutos','sum')).reset_index()
            jugadores_validos = df_squad_totals[df_squad_totals['Minutos_tot'] >= 300]['Jugador'].tolist()
            df_ref = df_raw[df_raw['Jugador'].isin(jugadores_validos)]
            maximos_equipo = []
            for m in metrics_radar:
                if m in df_ref.columns:
                    max_val = df_ref.groupby('Jugador')[m].sum().max()
                    maximos_equipo.append(max(max_val, 1))
                else:
                    maximos_equipo.append(1)
            valores_norm = [min((v/m*100) if m > 0 else 0, 100) for v, m in zip(totales_jugador, maximos_equipo)]
            fig_radar = go.Figure(data=go.Scatterpolar(r=valores_norm+[valores_norm[0]], theta=labels_radar+[labels_radar[0]], fill='toself', fillcolor='rgba(208,2,27,0.2)', line=dict(color='#D0021B', width=3), marker=dict(color='#D0021B', size=8), hovertemplate="%{theta}: %{r:.1f}%<extra></extra>"))
            fig_radar.update_layout(polar=dict(bgcolor="rgba(249,250,251,1)", radialaxis=dict(visible=True, range=[0,100], showticklabels=True, tickvals=[25,50,75,100], ticktext=['25%','50%','75%','100%'], tickfont=dict(size=9,color='#9CA3AF',family='Rajdhani'), gridcolor="#E5E7EB", linecolor="#E5E7EB"), angularaxis=dict(tickfont=dict(size=13,family="Rajdhani",color="#374151"), gridcolor="#E5E7EB", linecolor="#E5E7EB")), showlegend=False, paper_bgcolor="rgba(0,0,0,0)", margin=dict(l=40,r=40,t=20,b=20), height=380)
            st.plotly_chart(fig_radar, use_container_width=True)
        with c_metrics:
            st.markdown(f"<div class='section-title'>📋 TEMPORADA {temporada_sel}</div>", unsafe_allow_html=True)
            st.metric("Promedio SofaScore", f"{prom:.2f}")
            st.metric("Minutos Jugados", min_tot)
            st.metric("Goles", int(df_j['Goles'].sum()))
            st.metric("Asistencias", int(df_j['Asistencias'].sum()))
            st.metric("Participación en Goles", int(df_j['Goles'].sum() + df_j['Asistencias'].sum()))
            q_col = 'Quites (Tackles)' if 'Quites (Tackles)' in df_j.columns else 'Quites'
            recup = int(df_j[q_col].sum() + df_j['Intercepciones'].sum()) if q_col in df_j.columns else 0
            st.metric("Recuperaciones Totales", recup)

# ─── ESTADÍSTICAS DE EQUIPO ───────────────────────────────────────────────────
elif menu == "Estadísticas de Equipo":
    page_header("⚖️", "ESTADÍSTICAS DE EQUIPO", "Comparativa por partido")
    hojas = df_raw.drop_duplicates('Partido')[['Partido','Hoja_Original']].set_index('Partido').to_dict()['Hoja_Original']
    partido = st.selectbox("Seleccioná la fecha:", list(hojas.keys()))
    mostrar_marcador(EXCEL_ACTUAL, hojas[partido])
    df_equipo = extraer_estadisticas_equipo(str(EXCEL_ACTUAL), hojas[partido])
    if not df_equipo.empty:
        try:
            cols = df_equipo.columns.tolist()
            metrica_col = cols[0]; local_col = cols[1]; rival_col = cols[2]
            df_equipo[local_col] = pd.to_numeric(df_equipo[local_col].astype(str).str.replace('%','',regex=False), errors='coerce')
            df_equipo[rival_col] = pd.to_numeric(df_equipo[rival_col].astype(str).str.replace('%','',regex=False), errors='coerce')
            df_num = df_equipo.dropna(subset=[local_col, rival_col])
            if not df_num.empty:
                fig_eq = go.Figure()
                fig_eq.add_trace(go.Bar(name=str(local_col), x=df_num[metrica_col], y=df_num[local_col], marker_color='#D0021B', hovertemplate="%{x}<br>"+str(local_col)+": %{y}<extra></extra>"))
                fig_eq.add_trace(go.Bar(name=str(rival_col), x=df_num[metrica_col], y=df_num[rival_col], marker_color='#374151', hovertemplate="%{x}<br>"+str(rival_col)+": %{y}<extra></extra>"))
                apply_plotly_style(fig_eq, title="COMPARATIVA DE MÉTRICAS")
                fig_eq.update_layout(barmode='group', height=360, xaxis=dict(tickangle=-25), legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
                st.plotly_chart(fig_eq, use_container_width=True)
        except Exception:
            pass
        st.dataframe(df_equipo, hide_index=True, use_container_width=True)

# ─── ESTADÍSTICAS INDIVIDUALES ────────────────────────────────────────────────
elif menu == "Estadísticas Individuales":
    page_header("👤", "ESTADÍSTICAS INDIVIDUALES", "Top 7 por categoría")
    def extraer_exitosos(valor):
        try:
            return int(str(valor).replace("'","").split('/')[0]) if isinstance(valor,(str,int,float)) else 0
        except:
            return 0
    partido_sel = st.selectbox("Seleccioná la fecha:", df_raw['Partido'].unique())
    hoja_orig = df_raw[df_raw['Partido'] == partido_sel]['Hoja_Original'].iloc[0]
    mostrar_marcador(EXCEL_ACTUAL, hoja_orig)
    df_p = df_raw[df_raw['Partido'] == partido_sel].copy()
    if 'Pases (Comp/Tot)'   in df_p.columns: df_p['Pases Completados'] = df_p['Pases (Comp/Tot)'].apply(extraer_exitosos)
    if 'Duelos (Gan/Tot)'   in df_p.columns: df_p['Duelos Ganados']     = df_p['Duelos (Gan/Tot)'].apply(extraer_exitosos)
    if 'Regates (Exit/Tot)' in df_p.columns: df_p['Regates Exitosos']   = df_p['Regates (Exit/Tot)'].apply(extraer_exitosos)
    if 'Quites (Tackles)'   in df_p.columns: df_p = df_p.rename(columns={'Quites (Tackles)': 'Quites'})
    cols_num_ind = ['Efectividad Pases','Efectividad Duelos','Efectividad Regates','Tiros al Arco','Tiros Totales','Pases Clave','Intercepciones']
    for c in cols_num_ind:
        if c in df_p.columns:
            df_p[c] = pd.to_numeric(df_p[c], errors='coerce').fillna(0)
    categorias_ind = {
        "⭐ Nota SofaScore":   ("Nota SofaScore",    ['Jugador','Nota SofaScore']),
        "🛡️ Quites":           ("Quites",            ['Jugador','Quites']),
        "🛑 Intercepciones":   ("Intercepciones",    ['Jugador','Intercepciones']),
        "⚔️ Duelos Ganados":   ("Duelos Ganados",    ['Jugador','Duelos Ganados','Efectividad Duelos']),
        "🎯 Pases Completados":("Pases Completados", ['Jugador','Pases Completados','Efectividad Pases']),
        "🔑 Pases Clave":      ("Pases Clave",       ['Jugador','Pases Clave']),
        "⚡ Regates Exitosos": ("Regates Exitosos",  ['Jugador','Regates Exitosos']),
        "👟 Tiros al Arco":    ("Tiros al Arco",     ['Jugador','Tiros al Arco']),
    }
    cols_grid = st.columns(2)
    for i, (label, (sort_col, show_cols)) in enumerate(categorias_ind.items()):
        show_cols_existentes = [c for c in show_cols if c in df_p.columns]
        if sort_col not in df_p.columns:
            continue
        with cols_grid[i % 2]:
            st.markdown(f"<div class='section-title' style='font-size:20px;'>{label}</div>", unsafe_allow_html=True)
            df_top = df_p.nlargest(7, sort_col)[show_cols_existentes].reset_index(drop=True)
            df_top.index = df_top.index + 1
            st.dataframe(df_top, hide_index=False, use_container_width=True)
            st.markdown("<br>", unsafe_allow_html=True)

# ─── PARADO TÁCTICO ───────────────────────────────────────────────────────────
elif menu == "Parado Táctico":
    page_header("📋", "PARADO TÁCTICO", "Formación del partido")
    hojas = df_raw.drop_duplicates('Partido')[['Partido','Hoja_Original']].set_index('Partido').to_dict()['Hoja_Original']
    partido = st.selectbox("Fecha:", list(hojas.keys()))
    mostrar_marcador(EXCEL_ACTUAL, hojas[partido])
    img = extraer_imagen_incrustada(str(EXCEL_ACTUAL), hojas[partido], 0)
    if img:
        st.image(img, use_container_width=True)
    else:
        st.info("No hay imagen táctica disponible para este partido.")

# ─── MAPA DE TIROS ────────────────────────────────────────────────────────────
elif menu == "Mapa de Tiros":
    page_header("🎯", "MAPA DE TIROS", "Distribución de disparos por equipo")
    hojas = df_raw.drop_duplicates('Partido')[['Partido','Hoja_Original']].set_index('Partido').to_dict()['Hoja_Original']
    partido = st.selectbox("Fecha:", list(hojas.keys()))
    mostrar_marcador(EXCEL_ACTUAL, hojas[partido])
    c1, c2 = st.columns(2)
    with c1:
        img_r = extraer_imagen_incrustada(str(EXCEL_ACTUAL), hojas[partido], 1)
        if img_r:
            st.markdown("<div class='section-title section-title-red'>🔴 RIVER PLATE</div>", unsafe_allow_html=True)
            st.image(img_r, use_container_width=True)
        else:
            st.info("Sin mapa de tiros de River.")
    with c2:
        img_v = extraer_imagen_incrustada(str(EXCEL_ACTUAL), hojas[partido], 2)
        if img_v:
            st.markdown("<div class='section-title'>⚫ RIVAL</div>", unsafe_allow_html=True)
            st.image(img_v, use_container_width=True)
        else:
            st.info("Sin mapa de tiros del rival.")

# ─── CARA A CARA ──────────────────────────────────────────────────────────────
elif menu == "Cara a Cara":
    page_header("⚔️", "CARA A CARA", "Comparación de perfiles tácticos P90")

    @st.cache_data
    def cargar_datos_para_comparacion():
        todos_datos = []
        for anio, ruta in temporadas_dict.items():
            df, estado = cargar_datos_completos(ruta)
            if estado == "OK" and not df.empty:
                df['Temporada'] = anio
                todos_datos.append(df)
        if todos_datos:
            return pd.concat(todos_datos, ignore_index=True)
        return pd.DataFrame()

    df_todas_temporadas = cargar_datos_para_comparacion()
    if df_todas_temporadas.empty:
        st.error("No se pudieron cargar los datos de las temporadas.")
        st.stop()

    st.markdown("<div class='info-box'>Compará dos jugadores de cualquier temporada. Las estadísticas se normalizan por 90 minutos para comparaciones justas.</div>", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""<div class='player-header player-header-red'>🔴 JUGADOR 1</div>""", unsafe_allow_html=True)
        t_a = st.selectbox("Temporada:", anios_disponibles, key="ta")
        jugadores_a = sorted(df_todas_temporadas[df_todas_temporadas['Temporada'] == t_a]['Jugador'].unique())
        j_a = st.selectbox("Jugador:", jugadores_a, key="ja")
    with c2:
        st.markdown("""<div class='player-header player-header-gray'>⚪ JUGADOR 2</div>""", unsafe_allow_html=True)
        t_b = st.selectbox("Temporada:", anios_disponibles, key="tb")
        jugadores_b = sorted(df_todas_temporadas[df_todas_temporadas['Temporada'] == t_b]['Jugador'].unique())
        j_b = st.selectbox("Jugador:", jugadores_b, key="jb")

    def extraer_duelos(valor):
        try:
            if isinstance(valor, str):
                return int(valor.replace("'","").split('/')[0])
            return int(valor)
        except:
            return 0

    def get_stats_p90(df_source, name, temporada):
        data = df_source[(df_source['Jugador'] == name) & (df_source['Temporada'] == temporada)].copy()
        mins = data['Minutos'].sum()
        if mins == 0:
            return None
        q_col = 'Quites (Tackles)' if 'Quites (Tackles)' in data.columns else 'Quites'
        duelos  = data['Duelos (Gan/Tot)'].apply(extraer_duelos).sum()   if 'Duelos (Gan/Tot)'   in data.columns else 0
        regates = data['Regates (Exit/Tot)'].apply(extraer_duelos).sum() if 'Regates (Exit/Tot)' in data.columns else 0
        efect_pases = data['Efectividad Pases'].replace(0, np.nan).mean()
        efect_pases = efect_pases if not pd.isna(efect_pases) else 0
        return {
            'Mins': mins, 'Partidos': data['Partido'].nunique(),
            'Nota': data['Nota SofaScore'].mean(),
            'Goles': (data['Goles'].sum() / mins * 90),
            'Asist': (data['Asistencias'].sum() / mins * 90),
            'KP':    (data['Pases Clave'].sum() / mins * 90) if 'Pases Clave' in data.columns else 0,
            'Efect_Pases': efect_pases,
            'Regates': (regates / mins * 90),
            'Duelos':  (duelos  / mins * 90),
            'Quites':  (data[q_col].sum() / mins * 90) if q_col in data.columns else 0,
            'Inter':   (data['Intercepciones'].sum() / mins * 90),
        }

    s_a = get_stats_p90(df_todas_temporadas, j_a, t_a) if j_a else None
    s_b = get_stats_p90(df_todas_temporadas, j_b, t_b) if j_b else None

    if s_a and s_b:
        st.markdown("<hr style='margin:24px 0;'>", unsafe_allow_html=True)
        st.markdown("<div class='section-title'>📊 COMPARACIÓN GENERAL</div>", unsafe_allow_html=True)
        col_m1,col_m2,col_m3,col_m4,col_m5 = st.columns(5)
        def delta_fmt(val_b, val_a, fmt=".2f"):
            d = val_b - val_a
            return f"+{d:{fmt}}" if d > 0 else f"{d:{fmt}}"
        with col_m1:
            st.metric(f"Nota — {j_a[:10]}", f"{s_a['Nota']:.2f}")
            st.metric(f"Nota — {j_b[:10]}", f"{s_b['Nota']:.2f}", delta=delta_fmt(s_b['Nota'], s_a['Nota']))
        with col_m2:
            st.metric("Minutos", int(s_a['Mins']))
            st.metric("Minutos ", int(s_b['Mins']), delta=int(s_b['Mins']-s_a['Mins']))
        with col_m3:
            st.metric("Goles P90", f"{s_a['Goles']:.2f}")
            st.metric("Goles P90 ", f"{s_b['Goles']:.2f}", delta=delta_fmt(s_b['Goles'], s_a['Goles']))
        with col_m4:
            st.metric("Asist P90", f"{s_a['Asist']:.2f}")
            st.metric("Asist P90 ", f"{s_b['Asist']:.2f}", delta=delta_fmt(s_b['Asist'], s_a['Asist']))
        with col_m5:
            st.metric("Pases Clave P90", f"{s_a['KP']:.2f}")
            st.metric("Pases Clave P90 ", f"{s_b['KP']:.2f}", delta=delta_fmt(s_b['KP'], s_a['KP']))

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div class='section-title'>🛡️ PERFILES TÁCTICOS</div>", unsafe_allow_html=True)

        mets = ['Goles','Asist','KP','Efect_Pases','Regates','Duelos','Quites','Inter']
        labs = ['Goles P90','Asist P90','Pases Clave P90','Efect. Pases %','Regates P90','Duelos Gan P90','Quites P90','Intercep P90']

        def get_maximos_temporada(df, temporada):
            df_t = df[df['Temporada'] == temporada].copy()
            if df_t.empty:
                return [1.0]*8
            q_c = 'Quites (Tackles)' if 'Quites (Tackles)' in df_t.columns else 'Quites'
            d_tot = df_t['Duelos (Gan/Tot)'].apply(extraer_duelos)   if 'Duelos (Gan/Tot)'   in df_t.columns else pd.Series(0, index=df_t.index)
            r_tot = df_t['Regates (Exit/Tot)'].apply(extraer_duelos) if 'Regates (Exit/Tot)' in df_t.columns else pd.Series(0, index=df_t.index)
            df_t = df_t.copy()
            df_t['_duelos']  = d_tot.values
            df_t['_regates'] = r_tot.values
            agg_cols = {'Goles':'sum','Asistencias':'sum','Intercepciones':'sum','Minutos':'sum','Efectividad Pases':'mean','_duelos':'sum','_regates':'sum'}
            if 'Pases Clave' in df_t.columns:
                agg_cols['Pases Clave'] = 'sum'
            tg = df_t.groupby('Jugador').agg(agg_cols)
            if q_c in df_t.columns:
                tg[q_c] = df_t.groupby('Jugador')[q_c].sum()
            else:
                tg[q_c] = 0
            tg_valid = tg[tg['Minutos'] >= 300] if len(tg[tg['Minutos'] >= 300]) > 0 else tg
            mins_safe = tg_valid['Minutos'].replace(0, 1)
            pk_max = ((tg_valid['Pases Clave']/mins_safe)*90).fillna(0).max() if 'Pases Clave' in tg_valid.columns else 0.01
            return [
                ((tg_valid['Goles']/mins_safe)*90).fillna(0).max(),
                ((tg_valid['Asistencias']/mins_safe)*90).fillna(0).max(),
                pk_max,
                tg_valid['Efectividad Pases'].replace(0,np.nan).max() or 95.0,
                ((tg_valid['_regates']/mins_safe)*90).fillna(0).max(),
                ((tg_valid['_duelos']/mins_safe)*90).fillna(0).max(),
                ((tg_valid[q_c]/mins_safe)*90).fillna(0).max(),
                ((tg_valid['Intercepciones']/mins_safe)*90).fillna(0).max(),
            ]

        mx_a = get_maximos_temporada(df_todas_temporadas, t_a)
        mx_b = get_maximos_temporada(df_todas_temporadas, t_b)
        mx_global = [max(a,b,0.01) for a,b in zip(mx_a,mx_b)]

        def norm_vals(stats, maximos):
            vals = [stats[m] for m in mets]
            normed = [min((v/m*100) if m > 0 else 0, 100) for v,m in zip(vals,maximos)]
            return normed, vals

        vals_j1_norm, vals_j1_raw = norm_vals(s_a, mx_global)
        vals_j2_norm, vals_j2_raw = norm_vals(s_b, mx_global)

        def fmt_hover_val(val, idx):
            return f"{val:.1f}%" if idx == 3 else f"{val:.3f}"
        text_j1 = [f"{labs[i]}: {fmt_hover_val(vals_j1_raw[i],i)}" for i in range(len(labs))] + [f"{labs[0]}: {fmt_hover_val(vals_j1_raw[0],0)}"]
        text_j2 = [f"{labs[i]}: {fmt_hover_val(vals_j2_raw[i],i)}" for i in range(len(labs))] + [f"{labs[0]}: {fmt_hover_val(vals_j2_raw[0],0)}"]

        fig_radar = go.Figure()
        fig_radar.add_trace(go.Scatterpolar(r=vals_j1_norm+[vals_j1_norm[0]], theta=labs+[labs[0]], fill='toself', fillcolor='rgba(208,2,27,0.2)', line=dict(color='#D0021B',width=3), marker=dict(color='#D0021B',size=7), name=f"{j_a} ({t_a})", hoverinfo='text', text=text_j1))
        fig_radar.add_trace(go.Scatterpolar(r=vals_j2_norm+[vals_j2_norm[0]], theta=labs+[labs[0]], fill='toself', fillcolor='rgba(55,65,81,0.18)', line=dict(color='#374151',width=3), marker=dict(color='#374151',size=7), name=f"{j_b} ({t_b})", hoverinfo='text', text=text_j2))
        fig_radar.update_layout(polar=dict(bgcolor="rgba(249,250,251,1)", radialaxis=dict(visible=True,range=[0,100],showticklabels=True,tickvals=[25,50,75,100],ticktext=['25%','50%','75%','100%'],tickfont=dict(size=9,color='#9CA3AF',family='Rajdhani'),gridcolor="#E5E7EB",linecolor="#E5E7EB"), angularaxis=dict(tickfont=dict(size=12,family="Rajdhani",color="#374151"),gridcolor="#E5E7EB",linecolor="#E5E7EB")), showlegend=True, legend=dict(orientation="h",yanchor="bottom",y=-0.18,xanchor="center",x=0.5,font=dict(size=14,family="Rajdhani",color="#111827"),bgcolor="rgba(255,255,255,0.9)",bordercolor="#E5E7EB",borderwidth=1), paper_bgcolor="rgba(0,0,0,0)", margin=dict(l=80,r=80,t=20,b=80), height=560)
        st.plotly_chart(fig_radar, use_container_width=True)

        st.markdown("<div class='section-title'>📋 TABLA DETALLADA</div>", unsafe_allow_html=True)
        datos_tabla = pd.DataFrame({
            'Métrica': ['Partidos','Minutos','Nota SofaScore','Goles (P90)','Asistencias (P90)','Pases Clave (P90)','Efect. Pases %','Regates (P90)','Duelos Ganados (P90)','Quites (P90)','Intercepciones (P90)'],
            f'{j_a} ({t_a})': [int(s_a['Partidos']),int(s_a['Mins']),f"{s_a['Nota']:.2f}",f"{s_a['Goles']:.3f}",f"{s_a['Asist']:.3f}",f"{s_a['KP']:.3f}",f"{s_a['Efect_Pases']:.1f}%",f"{s_a['Regates']:.3f}",f"{s_a['Duelos']:.3f}",f"{s_a['Quites']:.3f}",f"{s_a['Inter']:.3f}"],
            f'{j_b} ({t_b})': [int(s_b['Partidos']),int(s_b['Mins']),f"{s_b['Nota']:.2f}",f"{s_b['Goles']:.3f}",f"{s_b['Asist']:.3f}",f"{s_b['KP']:.3f}",f"{s_b['Efect_Pases']:.1f}%",f"{s_b['Regates']:.3f}",f"{s_b['Duelos']:.3f}",f"{s_b['Quites']:.3f}",f"{s_b['Inter']:.3f}"],
        })
        st.dataframe(datos_tabla, hide_index=True, use_container_width=True)

# ─── HISTORIAL GENERAL ────────────────────────────────────────────────────────
elif menu == "Historial General":
    page_header("🌍", "HISTORIAL GENERAL", "Historial histórico vs todos los rivales registrados")
    condicion_sel = st.radio("Condición:", ["Total","Local","Visitante"], horizontal=True)
    st.markdown("<br>", unsafe_allow_html=True)
    df_historial = generar_historial_completo(condicion_sel)
    if df_historial.empty:
        st.info(f"No hay partidos registrados bajo la condición '{condicion_sel}'.")
    else:
        total_pj = df_historial['PJ'].sum(); total_pg = df_historial['PG'].sum()
        total_pe = df_historial['PE'].sum(); total_pp = df_historial['PP'].sum()
        efectividad = (total_pg*3+total_pe)/(total_pj*3)*100 if total_pj > 0 else 0
        c1,c2,c3,c4,c5 = st.columns(5)
        c1.metric("Total Partidos Registrados", int(total_pj)); c2.metric("Victorias Históricas", int(total_pg))
        c3.metric("Empates Históricos", int(total_pe)); c4.metric("Derrotas Históricas", int(total_pp))
        c5.metric("Efectividad Histórica", f"{efectividad:.1f}%")
        st.markdown("<br>", unsafe_allow_html=True)
        c_graf, c_tabla = st.columns([1.5, 1])
        with c_graf:
            st.markdown("<div class='section-title'>📊 RESULTADOS HISTÓRICOS POR RIVAL</div>", unsafe_allow_html=True)
            df_hist_graf = df_historial.sort_values(by='PJ', ascending=True)
            fig_hist = go.Figure()
            fig_hist.add_trace(go.Bar(y=df_hist_graf['Rival'], x=df_hist_graf['PG'], name='Ganados', orientation='h', marker=dict(color='#22C55E'), text=df_hist_graf['PG'].replace(0,''), textposition='inside', textfont=dict(color='white',family='Bebas Neue')))
            fig_hist.add_trace(go.Bar(y=df_hist_graf['Rival'], x=df_hist_graf['PE'], name='Empatados', orientation='h', marker=dict(color='#9CA3AF'), text=df_hist_graf['PE'].replace(0,''), textposition='inside', textfont=dict(color='white',family='Bebas Neue')))
            fig_hist.add_trace(go.Bar(y=df_hist_graf['Rival'], x=df_hist_graf['PP'], name='Perdidos', orientation='h', marker=dict(color='#EF4444'), text=df_hist_graf['PP'].replace(0,''), textposition='inside', textfont=dict(color='white',family='Bebas Neue')))
            apply_plotly_style(fig_hist, xaxis_title="Cantidad de Partidos", yaxis_title="")
            fig_hist.update_layout(barmode='stack', height=max(400,len(df_hist_graf)*35), legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1))
            st.plotly_chart(fig_hist, use_container_width=True)
            st.markdown("<div class='info-box'>💡 Los partidos definidos por penales se contabilizan estadísticamente como <b>Empate</b>.</div>", unsafe_allow_html=True)
        with c_tabla:
            st.markdown("<div class='section-title'>📋 TABLA DETALLADA</div>", unsafe_allow_html=True)
            st.dataframe(df_historial[['Rival','PJ','PG','PE','PP','GF','GC','DIF']], hide_index=True, use_container_width=True, height=800)

# ── FOOTER ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class='footer'>
    Data CARP · Club Atlético River Plate · Análisis de Rendimiento
</div>
""", unsafe_allow_html=True)
