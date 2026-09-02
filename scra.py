import tls_client
import pandas as pd
import matplotlib.pyplot as plt
from mplsoccer import Pitch
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
import os
import traceback
from datetime import datetime

# =========================================================
# 1. CONFIGURACIÓN
# =========================================================
EVENT_ID = 16671588

# RUTA CORREGIDA: Se usan rutas relativas para que funcione perfecto en Linux
RUTA_EXCEL = "Base_Datos_River_2026.xlsx"
CARPETA_TRABAJO = "." 

# =========================================================
# HEADERS SÚPER REFORZADOS (Para evadir Cloudflare/SofaScore)
# =========================================================
# Cambiamos a una versión más nueva de Chrome y activamos extensiones aleatorias
session = tls_client.Session(
    client_identifier="chrome_122", 
    random_tls_extension_order=True
)

headers = {
    "Accept": "*/*",
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
    "Origin": "https://www.sofascore.com",
    "Referer": "https://www.sofascore.com/",
    "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": '"Windows"',
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-site",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
}

ESTADISTICAS_CLAVE = {
    "Ball possession": "Posesión de balón",
    "Expected goals": "Goles esperados (xG)",
    "Total shots": "Tiros totales",
    "Shots on target": "Tiros al arco",
    "Shots off target": "Tiros afuera",
    "Blocked shots": "Tiros bloqueados",
    "Corner kicks": "Córners",
    "Offsides": "Fueras de juego",
    "Fouls": "Faltas",
    "Yellow cards": "Tarjetas amarillas",
    "Red cards": "Tarjetas rojas",
    "Passes": "Pases totales",
    "Accurate passes": "Pases precisos",
    "Accurate long balls": "Balones largos precisos",
    "Accurate crosses": "Centros precisos",
    "Goalkeeper saves": "Atajadas del arquero",
    "Tackles": "Quites",
    "Interceptions": "Intercepciones",
    "Clearances": "Despejes"
}

def formato_fraccion(acertados, totales): return f"'{acertados}/{totales}"
def calcular_porcentaje(acertados, totales): return int((acertados / totales) * 100) if totales > 0 else 0

def extraer_numero(valor):
    if isinstance(valor, str):
        valor = valor.replace('%', '')
    try: return float(valor)
    except: return 0.0

def ejecutar_reporte_exacto(event_id):
    print(f"--- Generando Reporte Exacto (ID: {event_id}) ---")

    try:
        # 2. INFO DEL PARTIDO
        print("Obteniendo datos del servidor...")
        response = session.get(f"https://api.sofascore.com/api/v1/event/{event_id}", headers=headers)
        
        if response.status_code != 200:
            print(f"❌ Error HTTP {response.status_code}: Posible bloqueo de SofaScore.")
            return
            
        res_info = response.json()
        
        if 'event' not in res_info:
            print("❌ Error: SofaScore no devolvió los datos. Te detectaron como Bot.")
            return
            
        event_data = res_info.get('event', {})
        home_data = event_data.get('homeTeam') or event_data.get('home_team')
        away_data = event_data.get('awayTeam') or event_data.get('away_team')
        
        if not home_data or not away_data:
            print("❌ Error: No se encontró la información de los equipos en este ID.")
            return
            
        local_name = home_data['name']
        lado_river = 'home' if 'River Plate' in local_name else 'away'
        rival = away_data['name'] if lado_river == 'home' else local_name

        # --- ARMADO DEL NOMBRE DE LA HOJA ---
        ts = event_data.get('startTimestamp')
        fecha_partido_dt = datetime.fromtimestamp(ts) if ts else datetime.now()
        fecha_partido_str = fecha_partido_dt.strftime('%d-%m')
        
        condicion = "(L)" if lado_river == 'home' else "(V)"
        
        torneo_nombre = event_data.get('tournament', {}).get('name', '').upper()
        if 'TROFEO' in torneo_nombre: comp_abrev = 'TC'
        elif 'SUPERCOPA' in torneo_nombre: comp_abrev = 'SA'
        elif 'COPA DE LA LIGA' in torneo_nombre: comp_abrev = 'CDL'
        elif 'COPA ARGENTINA' in torneo_nombre: comp_abrev = 'CA'
        elif 'LIBERTADORES' in torneo_nombre: comp_abrev = 'LIB'
        elif 'WORLD CUP' in torneo_nombre or 'MUNDIAL' in torneo_nombre: comp_abrev = 'MDC'
        elif 'LIGA PROFESIONAL' in torneo_nombre or 'LIGA' in torneo_nombre: comp_abrev = 'LPF'
        else: comp_abrev = torneo_nombre[:3]

        ronda_info = event_data.get('roundInfo', {})
        ronda_num = ronda_info.get('round')
        ronda_name = str(ronda_info.get('name', '')).upper()
        
        f_str = ""
        if 'FINAL' in ronda_name and 'QUARTER' not in ronda_name and 'SEMI' not in ronda_name and '1/8' not in ronda_name and '1/16' not in ronda_name: f_str = "FINAL"
        elif 'SEMI' in ronda_name: f_str = "SEMI"
        elif 'QUARTER' in ronda_name or '1/4' in ronda_name: f_str = "4TOS"
        elif '1/8' in ronda_name or 'ROUND OF 16' in ronda_name: f_str = "8VOS"
        elif '1/16' in ronda_name: f_str = "16VOS"
        elif ronda_num: f_str = f"F{ronda_num}"
        else: f_str = "P"
        
        prefijo = f"{fecha_partido_str} {f_str} VS "
        sufijo = f" {condicion} {comp_abrev}"
        caracteres_disp = 31 - len(prefijo) - len(sufijo)
        rival_corto = rival.upper()[:max(3, caracteres_disp)].strip()
        nombre_hoja = f"{prefijo}{rival_corto}{sufijo}"[:31]

        goles_local = event_data.get('homeScore', {}).get('current', 0)
        goles_visitante = event_data.get('awayScore', {}).get('current', 0)

        # 3. ESTADÍSTICAS GENERALES DEL PARTIDO
        print("Obteniendo estadísticas generales del partido...")
        res_stats = session.get(f"https://api.sofascore.com/api/v1/event/{event_id}/statistics", headers=headers).json()
        
        stats_temporales = {}
        if 'statistics' in res_stats and len(res_stats['statistics']) > 0:
            stats_all = next((p for p in res_stats['statistics'] if p['period'] == 'ALL'), None)
            if stats_all:
                for grupo in stats_all['groups']:
                    for item in grupo['statisticsItems']:
                        nombre_original = item.get('name', '')
                        if nombre_original not in stats_temporales:
                            stats_temporales[nombre_original] = {
                                'home': item.get('home', '0'),
                                'away': item.get('away', '0')
                            }

        estadisticas_partido = []
        stats_num_river = {}
        stats_num_rival = {}

        for clave_ingles, nombre_espanol in ESTADISTICAS_CLAVE.items():
            if clave_ingles in stats_temporales:
                val_home = stats_temporales[clave_ingles]['home']
                val_away = stats_temporales[clave_ingles]['away']
                estadisticas_partido.append([nombre_espanol, val_home, val_away])
                
                if lado_river == 'home':
                    stats_num_river[nombre_espanol] = extraer_numero(val_home)
                    stats_num_rival[nombre_espanol] = extraer_numero(val_away)
                else:
                    stats_num_river[nombre_espanol] = extraer_numero(val_away)
                    stats_num_rival[nombre_espanol] = extraer_numero(val_home)

        estadisticas_partido.insert(0, ["Resultado", goles_local, goles_visitante])

        # PUNTOS ALTOS Y BAJOS
        puntos_altos = []
        puntos_bajos = []

        if stats_num_river.get("Posesión de balón", 0) > 55: puntos_altos.append("Dominio de la posesión de balón")
        elif stats_num_river.get("Posesión de balón", 0) < 45: puntos_bajos.append("Dificultad para retener la pelota")

        if stats_num_river.get("Tiros al arco", 0) >= 5: puntos_altos.append("Buen volumen de tiros al arco")
        elif stats_num_river.get("Tiros totales", 0) < 8: puntos_bajos.append("Poca generación de juego ofensivo")

        if stats_num_rival.get("Tiros al arco", 0) <= 2: puntos_altos.append("Solidez defensiva (Permitió pocos tiros)")
        elif stats_num_rival.get("Tiros al arco", 0) >= 5: puntos_bajos.append("El rival llegó con facilidad al arco")

        pases_totales = stats_num_river.get("Pases totales", 0)
        pases_precisos = stats_num_river.get("Pases precisos", 0)
        if pases_totales > 0:
            if (pases_precisos / pases_totales) > 0.85: puntos_altos.append("Alta precisión en el armado de juego")
            elif (pases_precisos / pases_totales) < 0.75: puntos_bajos.append("Imprecisión general en los pases")

        if stats_num_river.get("Tiros afuera", 0) > stats_num_river.get("Tiros al arco", 0) and stats_num_river.get("Tiros totales", 0) > 10:
            puntos_bajos.append("Baja efectividad en la definición")

        if stats_num_river.get("Faltas", 0) > 13: puntos_bajos.append("Exceso de faltas cometidas")
        if stats_num_river.get("Despejes", 0) > 20 and stats_num_rival.get("Posesión de balón", 0) > 50:
            puntos_bajos.append("El equipo jugó sometido demasiado tiempo")

        if not puntos_altos: puntos_altos.append("Sin aspectos destacados")
        if not puntos_bajos: puntos_bajos.append("Sin aspectos negativos evidentes")

        # 4. ESTADÍSTICAS JUGADORES
        print("Procesando jugadores de River Plate...")
        res_lineups = session.get(f"https://api.sofascore.com/api/v1/event/{event_id}/lineups", headers=headers).json()
        jugadores_river = res_lineups[lado_river]['players']
        
        orden_pos = {'G': (1, 'Arquero'), 'D': (2, 'Defensor'), 'M': (3, 'Mediocampista'), 'F': (4, 'Delantero')}
        filas_excel, dicc_plantel = [], {}

        for j in jugadores_river:
            s = j.get('statistics', {})
            minutos = s.get('minutesPlayed', 0)
            if minutos == 0: continue
            
            p_id = j['player']['id']
            dicc_plantel[p_id] = {
                'nombre': j['player'].get('shortName', '').split()[-1].upper(),
                'numero': j.get('shirtNumber', ''),
                'es_titular': not j.get('substitute', False)
            }

            p_acc, p_tot = s.get('accuratePass', 0), s.get('totalPass', 0)
            d_gan, d_tot = s.get('duelWon', 0), s.get('duelWon', 0) + s.get('duelLost', 0)
            c_acc, c_tot = s.get('accurateCross', 0), s.get('totalCross', 0)
            b_acc, b_tot = s.get('accurateLongBalls', 0), s.get('totalLongBalls', 0)
            r_acc, r_tot = s.get('wonContest', 0), s.get('totalContest', 0)
            
            t_arco = s.get('onTargetScoringAttempt', 0)
            t_afuera = s.get('shotOffTarget', 0)
            t_tot = t_arco + t_afuera + s.get('blockedScoringAttempt', 0)

            filas_excel.append({
                'Jugador': j['player']['name'], 
                'Posición': orden_pos.get(j['player'].get('position', 'M'), (5, 'Otro'))[1],
                'Orden': orden_pos.get(j['player'].get('position', 'M'), (5, 'Otro'))[0],
                'Minutos': minutos, 
                'Nota SofaScore': s.get('rating', 0.0),
                'Quites (Tackles)': s.get('totalTackle', 0), 
                'Intercepciones': s.get('interceptionWon', 0),
                'Despejes': s.get('totalClearance', 0),
                'Duelos (Gan/Tot)': formato_fraccion(d_gan, d_tot), 
                'Efectividad Duelos': calcular_porcentaje(d_gan, d_tot),
                'Pérdidas de posesión': s.get('possessionLostCtrl', 0),
                'Faltas Cometidas': s.get('fouls', 0), 
                'Faltas Recibidas': s.get('wasFouled', 0),
                'Asistencias': s.get('goalAssist', 0), 
                'Pases Clave': s.get('keyPass', 0),
                'Pases (Comp/Tot)': formato_fraccion(p_acc, p_tot), 
                'Efectividad Pases': calcular_porcentaje(p_acc, p_tot),
                'Centros (Comp/Tot)': formato_fraccion(c_acc, c_tot), 
                'Efectividad Centros': calcular_porcentaje(c_acc, c_tot),
                'Balones Largos (C/T)': formato_fraccion(b_acc, b_tot), 
                'Efectividad Balones': calcular_porcentaje(b_acc, b_tot),
                'Regates (Exit/Tot)': formato_fraccion(r_acc, r_tot), 
                'Efectividad Regates': calcular_porcentaje(r_acc, r_tot),
                'Goles': s.get('goals', 0), 
                'Tiros Totales': t_tot, 
                'Tiros al Arco': t_arco, 
                'Tiros Afuera': t_afuera
            })

        df = pd.DataFrame(filas_excel).sort_values(by=['Orden', 'Minutos'], ascending=[True, False]).drop(columns=['Orden'])

        # 5. IMÁGENES TÁCTICAS
        print("Generando gráficos tácticos...")
        pitch = Pitch(pitch_type='opta', pitch_color='#1a1a1a', line_color='#555555')
        fig1, ax1 = pitch.draw(figsize=(8, 5.5))
        fig1.patch.set_facecolor('#1a1a1a')
        
        res_pos = session.get(f"https://api.sofascore.com/api/v1/event/{event_id}/average-positions", headers=headers).json()
        for p in res_pos.get(lado_river, []):
            if 'player' not in p: continue
            id_j = p['player']['id']
            if id_j in dicc_plantel and dicc_plantel[id_j]['es_titular']:
                jug = dicc_plantel[id_j]
                ax1.scatter(p['averageX'], p['averageY'], s=500, color="#d32f2f", edgecolor="white", zorder=5)
                ax1.text(p['averageX'], p['averageY'], str(jug['numero']), color="white", fontsize=9, ha="center", va="center", fontweight="bold")
                ax1.text(p['averageX'], p['averageY'] + 4.5, jug['nombre'], color="white", fontsize=7, ha="center", fontweight='bold')

        plt.title("POSICIONES MEDIAS - RIVER PLATE", color="white", fontsize=12, pad=10)
        img_parado = os.path.join(CARPETA_TRABAJO, "parado.png")
        plt.savefig(img_parado, bbox_inches="tight", dpi=120, facecolor='#1a1a1a')
        plt.close(fig1)

        # 6. SHOTMAPS (RIVER Y RIVAL)
        fig2, ax2 = pitch.draw(figsize=(8, 5.5))
        fig2.patch.set_facecolor('#1a1a1a')
        
        fig3, ax3 = pitch.draw(figsize=(8, 5.5))
        fig3.patch.set_facecolor('#1a1a1a')
        
        res_shots = session.get(f"https://api.sofascore.com/api/v1/event/{event_id}/shotmap", headers=headers).json()
        
        for s in res_shots.get('shotmap', []):
            x, y = s['playerCoordinates']['x'], s['playerCoordinates']['y']
            nombre = s['player']['shortName'].split()[-1].upper()
            tipo = s['shotType']
            
            if tipo == 'goal': color, marker, size = "#FFD700", "*", 250 
            elif tipo == 'save': color, marker, size = "#3498db", "s", 100 
            elif tipo == 'block': color, marker, size = "#9b59b6", "^", 100 
            else: color, marker, size = "#e74c3c", "o", 100 
            
            if s['isHome'] == (lado_river == 'home'):
                ax2.scatter(x, y, s=size, color=color, marker=marker, edgecolor="white", alpha=0.9, zorder=5)
                ax2.text(x, y + 3, nombre, color="white", fontsize=6, ha="center")
            else:
                ax3.scatter(x, y, s=size, color=color, marker=marker, edgecolor="white", alpha=0.9, zorder=5)
                ax3.text(x, y + 3, nombre, color="white", fontsize=6, ha="center")

        for ax in [ax2, ax3]:
            ax.scatter(5, 95, s=100, color="#FFD700", marker="*", edgecolor="white")
            ax.text(7, 94.5, "Gol", color="white", fontsize=8)
            ax.scatter(18, 95, s=60, color="#3498db", marker="s", edgecolor="white")
            ax.text(20, 94.5, "Atajado", color="white", fontsize=8)
            ax.scatter(38, 95, s=60, color="#9b59b6", marker="^", edgecolor="white")
            ax.text(40, 94.5, "Bloqueado", color="white", fontsize=8)
            ax.scatter(55, 95, s=60, color="#e74c3c", marker="o", edgecolor="white")
            ax.text(57, 94.5, "Afuera", color="white", fontsize=8)

        ax2.set_title("MAPA DE TIROS - RIVER PLATE", color="white", fontsize=12, pad=10)
        img_shots = os.path.join(CARPETA_TRABAJO, "shots.png")
        fig2.savefig(img_shots, bbox_inches="tight", dpi=120, facecolor='#1a1a1a')
        plt.close(fig2)

        ax3.set_title(f"MAPA DE TIROS - {rival.upper()}", color="white", fontsize=12, pad=10)
        img_shots_rival = os.path.join(CARPETA_TRABAJO, "shots_rival.png")
        fig3.savefig(img_shots_rival, bbox_inches="tight", dpi=120, facecolor='#1a1a1a')
        plt.close(fig3)

        # 7. GUARDADO EXCEL Y FORMATO
        print("Guardando en Excel y estructurando Dashboard...")
        if os.path.exists(RUTA_EXCEL):
            with pd.ExcelWriter(RUTA_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, index=False, sheet_name=nombre_hoja)
        else:
            df.to_excel(RUTA_EXCEL, index=False, sheet_name=nombre_hoja)

        wb = load_workbook(RUTA_EXCEL)
        ws = wb[nombre_hoja]
        ws.freeze_panes = 'B2'
        
        for idx in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16

        rojo = 'F8696B'
        amarillo = 'FFEB84'
        verde = '63BE7B'
        blanco = 'FFFFFF'

        regla_verde_bueno = ColorScaleRule(start_type='min', start_color=rojo, mid_type='percentile', mid_value=50, mid_color=amarillo, end_type='max', end_color=verde)
        regla_rojo_bueno = ColorScaleRule(start_type='min', start_color=verde, mid_type='percentile', mid_value=50, mid_color=amarillo, end_type='max', end_color=rojo)
        regla_solo_verde = ColorScaleRule(start_type='min', start_color=blanco, end_type='max', end_color=verde)
        regla_solo_rojo = ColorScaleRule(start_type='min', start_color=blanco, end_type='max', end_color=rojo)

        columnas_df = list(df.columns)
        ultima_fila_jugadores = len(df) + 1

        def aplicar_color(nombre_col, regla):
            if nombre_col in columnas_df:
                letra = get_column_letter(columnas_df.index(nombre_col) + 1)
                ws.conditional_formatting.add(f'{letra}2:{letra}{ultima_fila_jugadores}', regla)

        for col in ['Nota SofaScore', 'Efectividad Duelos', 'Efectividad Pases', 'Efectividad Centros', 'Efectividad Balones', 'Efectividad Regates']: aplicar_color(col, regla_verde_bueno)
        for col in ['Pérdidas de posesión']: aplicar_color(col, regla_rojo_bueno)
        for col in ['Minutos', 'Quites (Tackles)', 'Intercepciones', 'Despejes', 'Faltas Recibidas', 'Asistencias', 'Pases Clave', 'Goles', 'Tiros Totales', 'Tiros al Arco']: aplicar_color(col, regla_solo_verde)
        for col in ['Faltas Cometidas', 'Tiros Afuera']: aplicar_color(col, regla_solo_rojo)

        # --- BLOQUE INFERIOR: STATS GENERALES Y CONCLUSIONES ---
        fila_inicio_stats = ultima_fila_jugadores + 3
        col_stats = 2 
        
        ws.cell(row=fila_inicio_stats, column=col_stats, value="ESTADÍSTICAS DEL PARTIDO").font = Font(bold=True)
        ws.cell(row=fila_inicio_stats+1, column=col_stats, value="Métrica").font = Font(bold=True)
        ws.cell(row=fila_inicio_stats+1, column=col_stats+1, value=local_name).font = Font(bold=True)
        ws.cell(row=fila_inicio_stats+1, column=col_stats+2, value=away_data['name']).font = Font(bold=True)
        
        fila_actual = fila_inicio_stats + 2
        for stat in estadisticas_partido:
            ws.cell(row=fila_actual, column=col_stats, value=stat[0])
            ws.cell(row=fila_actual, column=col_stats+1, value=stat[1])
            ws.cell(row=fila_actual, column=col_stats+2, value=stat[2])
            fila_actual += 1

        col_altos = 5 
        col_bajos = 6 
        
        ws.cell(row=fila_inicio_stats, column=col_altos, value="PUNTOS ALTOS").font = Font(bold=True, color="008000")
        ws.cell(row=fila_inicio_stats, column=col_bajos, value="PUNTOS BAJOS").font = Font(bold=True, color="FF0000")
        
        for idx, punto in enumerate(puntos_altos):
            ws.cell(row=fila_inicio_stats + 1 + idx, column=col_altos, value=f"• {punto}")
            
        for idx, punto in enumerate(puntos_bajos):
            ws.cell(row=fila_inicio_stats + 1 + idx, column=col_bajos, value=f"• {punto}")

        ws.column_dimensions[get_column_letter(col_altos)].width = 38
        ws.column_dimensions[get_column_letter(col_bajos)].width = 38

        # --- IMÁGENES A LA DERECHA ---
        fila_img = fila_inicio_stats
        
        img1 = XLImage(img_parado); img1.width, img1.height = 500, 360 
        ws.add_image(img1, f'H{fila_img}') 
        
        img2 = XLImage(img_shots); img2.width, img2.height = 500, 360
        ws.add_image(img2, f'P{fila_img}') 

        img3 = XLImage(img_shots_rival); img3.width, img3.height = 500, 360
        ws.add_image(img3, f'H{fila_img + 20}') 

        wb.save(RUTA_EXCEL)
        wb.close()
        
        for f in [img_parado, img_shots, img_shots_rival]:
            if os.path.exists(f): os.remove(f)
            
        print(f"✔️ ¡Reporte generado con ÉXITO en la hoja '{nombre_hoja}'!")

    except Exception:
        traceback.print_exc()

ejecutar_reporte_exacto(EVENT_ID)